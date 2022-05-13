import openpyxl as xl
import datetime
thisyear = datetime.date.today().year

wareki_table = [
    {"name":"明治", "start":1868, "end":1912},
    {"name":"大正", "start":1912, "end":1926},
    {"name":"昭和", "start":1926, "end":1989},
    {"name":"平成", "start":1989, "end":2019},
    {"name":"令和", "start":2019, "end":thisyear}

]
def seireki_wareki(year):
    for w in wareki_table:
        if w["start"] <= year < w["end"]:
            y = str(year - w["start"] + 1) + "年"
            if y == "1年": y = "元年"
            return w["name"]+y
        elif year == w["end"] and year == thisyear:
            y = str(year - w["start"] + 1) + "年"
            return w["name"] + y
    return "不明"
book = xl.Workbook()
sheet = book.active

sheet["A1"] = "西暦"
sheet["B1"] = "和暦"


print(thisyear)

start_y = 1868
for i in range(1868, thisyear+1):
    sei = start_y + (i-1868)
    wa = seireki_wareki(sei)
    sheet.cell(row=(2+(i-1868)), column=1, value=str(sei)+"年")
    sheet.cell(row=(2+(i-1868)), column=2, value=wa)
    print(sei, "=", wa)
book.save("wareki.xlsx")