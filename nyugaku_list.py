import openpyxl as xl
import datetime
book = xl.Workbook()
sheet = book.active

base_year = datetime.date.today().year-10

for i in range(50):
    year = base_year - i
    s2 = year - 6
    s1 = year - 7
    sf = f'{s1}年4/2～{s2}年4/1に生まれた人'
    sheet.cell(i+1, 1, value=str(year)+"年度")
    sheet.cell(i+1, 2, value=sf)

book.save("nyugaku_list.xlsx")