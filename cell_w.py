import openpyxl as xl
book = xl.Workbook()
sheet = book.active

sheet["A1"] = "お腹がへった"
sheet.cell(row=2, column=1, value="もうすぐ休憩")
cell = sheet.cell(row=3,column=1)
cell.value = "もうちょっと"
book.save("cell_w.xlsx")

