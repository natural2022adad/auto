import openpyxl as xl
import datetime as dat
book = xl.load_workbook("wareki.xlsx")
sheet = book.active

it = sheet.iter_rows()

for row in it:
    r = []
    for cell in row:
        r.append(cell.value)
    print(r)

cell = sheet["U59"]
(row, col) = (cell.row, cell.column)
print(f'U59)=({row}, {col})')

cell = sheet.cell(row=2, column=3)
cdt = cell.coordinate
print(f"(2, 3)={cdt}")